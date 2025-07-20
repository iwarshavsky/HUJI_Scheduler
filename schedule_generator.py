import io
import json
import os
import struct
from datetime import datetime, timedelta
from urllib.parse import urlencode

from flask import Flask

from bs4 import BeautifulSoup
import argparse
from copy import copy, deepcopy
from types import SimpleNamespace
import timeit
import itertools
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
import pickle
from flask import session



# Custom headers
headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Accept-Language": "en-US,en;q=0.9,he;q=0.8",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "Content-Type": "application/x-www-form-urlencoded",
    "Host": "shnaton.huji.ac.il",
    "Origin": "https://shnaton.huji.ac.il",
    "Referer": "https://shnaton.huji.ac.il/index.php",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0",
    "sec-ch-ua": '"Not)A;Brand";v="8", "Chromium";v="138", "Microsoft Edge";v="138"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"'
}


def make_request(course_id, year):
    base_url = "https://shnaton.huji.ac.il/index.php"
    # Make the POST request
    data = {
        "year": year,
        "faculty": "0",
        "peula": "Simple",
        "option": "2",  # Milim beshem hakurs
        "word": "",
        "course": course_id
    }

    query_string = urlencode(data)

    # response = requests.post(url, data=data, headers=headers)  # This is a blocking call
    full_url = f"{base_url}?{query_string}"
    response = requests.get(full_url, headers=headers)
    if response.status_code != 200:
        print("Error. Received code " + str(response.status_code))
        exit()

    # # Save the HTML response
    # with open("response.html", "w", encoding="utf-8") as f:
    #     f.write(response.text)

    return response.text, full_url


# TODO pickle?

class CoursePool:
    def __init__(self):
        pass


class CourseBlock:
    def __init__(self, course_num, course_name, bs_el=None, data_dict=None):
        """
        Initialiaze a CourseBlock from raw HTML row - may include several events.
        We assume each row contains info of the same semester, same group.
        TIME_FINISH is not inclusive.
        :param bs_el: BeautifulSoup element
        """
        self.course_num = course_num
        self.course_name = course_name
        if bs_el:
            raw_keys = ["groups", "semester", "days", "hour", "lesson", "places", "note"]
            keys = ["groups", "semester", "days", "time_start", "time_finish", "lesson", "places", "note"]
            raw_attributes, self.warnings = self._extract_raw_attributes(bs_el, raw_keys)
            self._split_hour_column(raw_attributes)
            # Remove rows with no date/time/semester

            formatted_attributes = {key: list(map(lambda item: self.format(key, item), raw_attributes[key])) for key in
                                    keys}

            # formatted_attributes['time_start_slot'] = [self.convert_time_to_slot(t) for t in formatted_attributes['time_start']]
            # formatted_attributes['time_finish_slot'] = [self.convert_time_to_slot(t) for t in formatted_attributes['time_finish']]


            self.group = formatted_attributes.get("groups")[0]
            self.semester = formatted_attributes.get("semester")[0]
            # TODO - check for empty rows
            # TODO - custom events
            self.event_list = self._make_event_list(formatted_attributes,
                                                    keys=["days", "time_start", "time_finish", "lesson", "places",
                                                          "note"])

            self.lesson_type = formatted_attributes.get("lesson")[0]


        if data_dict:
            self.group = data_dict['group']
            self.semester = data_dict['semester']
            self.event_list = data_dict['event_list']
            self.lesson_type = data_dict['lesson_type']
            # TODO : should check validity of keys or am I OK with throwing an exception here?
        pass

    @staticmethod
    def _extract_raw_attributes(bs_el, attribute_lst):
        # return bs_el.find("div", {"class": "semester"}).div.string and warning if there was an error parsing
        attribute_dict = {}
        for attribute in attribute_lst:
            attribute_dict[attribute] = list(bs_el.find("div", {"class": attribute}).stripped_strings)
        warnings = []
        # Find max len of any key, pad the rest. Result is dict with list values of equal length
        max_len = max([len(attribute_data) for attribute_data in attribute_dict.values()])
        for attribute in attribute_lst:
            l = len(attribute_dict[attribute])
            if l != max_len:
                attribute_dict[attribute] += ['']*(max_len-l)
                warnings.append(attribute)
        return attribute_dict, warnings

    @staticmethod
    def _split_hour_column(raw_attributes):
        """
        Replace "hour" key with time_start and time_finish
        """
        dict_new_cols = {"time_start": [], "time_finish": []}
        for t_string in raw_attributes["hour"]:
            if t_string:
                time_finish, time_start = t_string.split("-")
                dict_new_cols["time_start"].append(time_start)
                dict_new_cols["time_finish"].append(time_finish)
        raw_attributes.pop("hour")

        raw_attributes["time_start"] = dict_new_cols["time_start"]
        raw_attributes["time_finish"] = dict_new_cols["time_finish"]

    @staticmethod
    def _make_event_list(raw_attribute_dict, keys):
        """
        Reorder the raw attribute dict so that the information isn't spread out between the different columns but
        is now divided into scheduled events
        :param raw_attribute_dict:
        :param keys:
        :return: array of scheduled events for this CourseBlock
        """
        # Make sure that there was no error parsing, all keys must have the same number of scheduled events
        if not all(len(raw_attribute_dict[k]) == len(raw_attribute_dict[keys[0]]) for k in keys):  # Invalid row
            print("Bad course")
            return []

        events = [{key: raw_attribute_dict[key][i] for key in keys}
                  for i in range(len(raw_attribute_dict[keys[0]]))]

        return events

    @staticmethod
    def format(key, value):
        try:
            if key == "semester":
                semester_mapping = {"סמסטר א": 0, "סמסטר ב": 1}
                return semester_mapping[value]
            elif key == "days":
                v = value.replace("יום ", "").replace("'", "")
                return ord(v) - 1488  # Convert Hebrew letter to corresponding number
            elif key in ["time_start", "time_finish"]:
                pass
                # t = datetime.strptime(value, "%-H:%M")
                # return t  # timedelta(hours=t.hour, minutes=t.minute)
                # return convert_time(value)
            elif key in "groups":
                v = value.replace("קבוצה (", "").replace(")", "")
                return v
                pass
            else:
                pass
        except:
            pass

        return value


class Course:
    """
    Multiple course blocks
    """

    def __init__(self, course_num, year, semester, pool_dict=None):
        self.course_num = course_num
        self.year = year
        self.semester = semester
        if pool_dict:
            self.pool_dict = pool_dict
        else:
            self.pool_dict = {}
            response, self.url = make_request(course_num, year)

            soup = BeautifulSoup(response, 'html.parser')

            COURSE_TABLE = "grid-container-wrapper"
            FIRST_ROW = "title-row"
            course_table = soup.find("div", COURSE_TABLE)
            soup.prettify()

            self.course_name = soup.select(".data-course-title")[0].text.strip()

            # Loop through children, first child has class="row title-row", should be ignored, the other have class="row"
            for row in course_table.find_all("div", recursive=False):
                if FIRST_ROW in row['class']:  # Ignore first row
                    continue
                session_group = CourseBlock(self.course_num, self.course_name, row)
                self.add_block(session_group)


            # print(row)
        # categorize Course Blocks by activity type - we must choose one of each activity type

    def add_block(self, courseBlock):
        if courseBlock.event_list and self.semester == courseBlock.semester:
            blocks_of_same_type = self.pool_dict.setdefault(courseBlock.lesson_type, [])
            blocks_of_same_type.append(courseBlock)

    def get_pools(self):
        return {(self.course_num, key): self.pool_dict[key] for key in self.pool_dict.keys()}


class Schedule:
    """
    Represents a 2D weekly agenda
    """
    SLOTS = 66
    DAYS = 6
    BASEHOUR = 8

    def __init__(self, data=None, blocks=None):

        if data and blocks:
            self.data = data
            self.blocks = blocks
        else:
            self.data = [None] * Schedule.SLOTS * Schedule.DAYS
            self.blocks = []

    def __getitem__(self, item):
        if isinstance(item, tuple):
            day, slot = item
            return self.data[day * Schedule.SLOTS + slot]
        if isinstance(item, int):
            day = item
            return self.data[day * Schedule.SLOTS:(day + 1) * Schedule.SLOTS]

    def __setitem__(self, key, value):
        day, slot = key
        self.data[day * Schedule.SLOTS + slot] = value

    def copy(self):
        return Schedule(self.data.copy(), self.blocks.copy())

    def to_dict(self):
        d = {i: self.blocks[i].__dict__ for i in range(len(self.blocks))}
        return d

    @staticmethod
    def convert_time(origin, dest_class, basehour=0):
        if isinstance(origin, dest_class):
            return origin
        if isinstance(origin, int) and dest_class is str:
            hour = (origin // 4) + basehour
            minutes = (origin % 4) * 15
            return f"{str(hour)}:{str(minutes)}"
        elif isinstance(origin, str) and dest_class is datetime:
            if not origin or ":" not in origin:
                return None
            hour_str, minute_str = origin.split(":")
            return datetime.strptime(f"{hour_str:0>2}:{minute_str}", "%H:%M")
        elif isinstance(origin, int) and dest_class is datetime:  # composition
            return Schedule.convert_time(Schedule.convert_time(origin, str, basehour), datetime)
        elif isinstance(origin, datetime) and dest_class is str:
            return origin.strftime("%H:%M")
        elif isinstance(origin, str) and dest_class is int:
            hour, minute = origin.split(":")
            time_val = int((int(hour) - basehour) * 4 + int(minute) / 15)
            return time_val
        elif isinstance(origin, datetime) and dest_class is int:
            return Schedule.convert_time(Schedule.convert_time(origin, str, basehour), int)

    def add_block(self, block: CourseBlock):
        for event in block.event_list:
            day = event["days"]
            for t in range(self.convert_time(event["time_start"], int, basehour=self.BASEHOUR),
                           self.convert_time(event["time_finish"], int, basehour=self.BASEHOUR)):
                if self[day, t]:
                    return None
                self[day, t] = str(block.course_num) + " " + str(block.group) + " " + block.lesson_type
        self.blocks.append(block)
        return self

    def get_properties(self):
        day_data = [self.get_day_info(i) for i in range(Schedule.DAYS)]
        return Schedule.Properties(
            max_day_length=max([day_data[i].day_length for i in range(Schedule.DAYS)]),
            min_day_length=min([day_data[i].day_length for i in range(Schedule.DAYS)]),
            free_days=sum([day_data[i].free_day for i in range(Schedule.DAYS)]),
            max_break_length=max([day_data[i].longest_break for i in range(Schedule.DAYS)]),
            globalStartTime=max([day_data[i].start_time for i in range(Schedule.DAYS)]),
            day_and_break_lengths=[{"day_length": day_data[i].day_length, "break_length": day_data[i].longest_break} for
                                   i in range(Schedule.DAYS)]
        )

    def get_day_info(self, day):
        return self.DayInfo(self[day])

    def get_blocks(self):
        return self.blocks


    class DayInfo:

        def __init__(self, data_arr):
            self.start_slot = self.get_start_slot(data_arr)
            self.finish_slot = self.get_finish_slot(data_arr)
            self.start_time = Schedule.convert_time(self.start_slot, datetime, Schedule.BASEHOUR)
            self.finish_time = Schedule.convert_time(self.finish_slot, datetime, Schedule.BASEHOUR)
            self.day_length = self.finish_time - self.start_time
            self.free_day = True if self.day_length else False
            self.longest_break = Schedule.convert_time(
                self.get_longest_break(data_arr, self.start_slot, self.finish_slot), datetime)

        @staticmethod
        def get_start_slot(arr):
            for i in range(Schedule.SLOTS):
                if arr[i]:
                    return i
            return 0

        @staticmethod
        def get_finish_slot(arr):
            for i in range(Schedule.SLOTS - 1, -1, -1):
                if arr[i]:
                    return i
            return 0

        @staticmethod
        def get_longest_break(arr, start_time, finish_time):
            longest_break = 0
            cur_break_len = 0
            for i in range(start_time, finish_time):
                if arr[i]:
                    if cur_break_len > 0:
                        longest_break = max(cur_break_len, longest_break)
                        cur_break_len = 0
                else:
                    cur_break_len += 1
            return longest_break

    class Properties:
        def __init__(self, max_day_length=None, min_day_length=None, free_days=None, max_break_length=None,
                     globalStartTime=None,
                     day_and_break_lengths=None, **kwargs):
            self.max_day_length = Schedule.convert_time(max_day_length, datetime)
            self.min_day_length = Schedule.convert_time(min_day_length, datetime)
            self.free_days = free_days
            self.max_break_length = Schedule.convert_time(max_break_length, datetime)
            self.globalStartTime = Schedule.convert_time(globalStartTime, datetime)
            self.day_and_break_lengths = day_and_break_lengths

    def to_excel(self):
        wb = load_workbook('static/template.xlsx')
        ws = wb["Results"]
        for block in self.blocks:
            for event in block.event_list:
                ws.cell(row=event['time_start'] + 2,
                        column=event['days'] + 1,
                        value='\n'.join([block.course_num, event['lesson'], event['places']]))
                ws.merge_cells(start_row=event['time_start'] + 2,
                               start_column=event['days'] + 1,
                               end_row=event['time_finish'] + 2,
                               end_column=event['days'] + 1)
        virtual_wb = io.BytesIO()
        wb.save(virtual_wb)
        wb.close()
        virtual_wb.seek(0)
        return virtual_wb

class Constraint(Schedule.Properties):
    def __init__(self, dayWithBreak_DayLength, dayWithBreak_BreakLength, **kwargs):
        super().__init__(**kwargs)
        self.dayWithBreak_DayLength = Schedule.convert_time(dayWithBreak_DayLength, datetime)
        self.dayWithBreak_BreakLength = Schedule.convert_time(dayWithBreak_BreakLength, datetime)

    def complies(self, schedule) -> bool:
        prop_other = schedule.get_properties()
        if self.max_day_length and not (prop_other.max_day_length <= self.max_day_length):
            return False
        if self.min_day_length and not (prop_other.min_day_length >= self.min_day_length):
            return False
        if self.free_days and not (prop_other.free_days >= self.free_days):
            return False
        if self.max_break_length and not (prop_other.max_break_length <= self.max_break_length):
            return False
        if self.globalStartTime and not (prop_other.globalStartTime >= self.globalStartTime):
            return False
        if self.dayWithBreak_DayLength and self.dayWithBreak_BreakLength:
            for d in prop_other.day_and_break_lengths:
                day_length = d["day_length"]
                break_length = d["break_length"]
                if day_length > self.dayWithBreak_DayLength and self.dayWithBreak_BreakLength < break_length:
                    return False

        return True


def scheduleGenerator(args):
    # if bool(args.dayWithBreak_DayLength) ^ bool(args.dayWithBreak_BreakLength):
    #     parser.error("dayWithBreak_DayLength and dayWithBreak_BreakLength must be given together")
    # print(args)

    # Convert
    constraint = Constraint(**args)

    # dayWithBreak_DayLength=, dayWithBreak_BreakLength=, max_day_length=, min_day_length=,
    # free_days=, max_break_length=, globalStartTime=, day_and_break_lengths=

    courses_dict = args['courses']

    # Reconstruct courses from dicts
    courses = [Course(course_num=course_dict['course'],
                      year=course_dict['year'],
                      semester=course_dict['semester'],
                      pool_dict={lesson_type: [CourseBlock(course_num=block_dict['course_num'],
                                                           course_name=block_dict['course_name'],
                                                           data_dict=block_dict)
                                               for block_dict in block_lst]
                                 for lesson_type, block_lst in course_dict['pool_dict'].items()})
               for course_dict in courses_dict.values()]

    pools = {k: v for course in courses for k, v in course.get_pools().items()}

    pools_keys = list(pools.keys())
    _pools = list(pools.values())
    # print(len(_pools))
    n = len(pools.keys())
    schedules = []

    def schedule_generator(i, schedule: Schedule):
        if i >= n:
            if constraint.complies(schedule):
                yield schedule
                return
        for block in _pools[i]:
            new_schedule = schedule.copy().add_block(block)
            if new_schedule:
                yield from schedule_generator(i + 1, new_schedule)

    yield from schedule_generator(0, schedule=Schedule())


if __name__ == "__main__":
    # GET COURSES FIRST
    with open("d", mode="rb") as f:
        d = pickle.load(f)
    res = list(scheduleGenerator(d))
    pass
    # {'courses': ['80131', '67101', '80181'], 'dayWithBreak_BreakLength': '', 'dayWithBreak_DayLength': '',
    #  'globalStartTime': '',
    #  'max_day_length': '', 'min_day_length': '', 'min_free_days': '', 'semester': 0, 'year': '2025'}

    # parser = argparse.ArgumentParser(
    #     description='Create a schedule')
    # parser.add_argument('--max_day_length', metavar='float', type=float, help='Maximal length of day', default=None)
    # parser.add_argument('--min_day_length', metavar='float', type=float, help='Minimal length of day', default=None)
    # parser.add_argument('--min_free_days', metavar='int', type=int, help='Minimal numbers of free days in a week', default=None)
    # parser.add_argument('--max_break_length', metavar='float', type=float, help='Days may not include breaks longer than this', default=None)
    # parser.add_argument('--dayWithBreak_DayLength', metavar='float', type=float, help='If day is longer than this number, then it must have a break of at least the next arg', default=None)
    # parser.add_argument('--dayWithBreak_BreakLength', metavar='float', type=float, help='If day is longer than the prev arg, then it must have a break of at least this', default=None)
    # parser.add_argument('--globalStartTime', metavar='float', type=float, help='Earliest time to start the day', default=None)
    # parser.add_argument('--year', metavar='int', type=int, help='What year')
    # parser.add_argument('--semester', metavar='int', type=int, help='What semester')
    #
    # args = parser.parse_args()
